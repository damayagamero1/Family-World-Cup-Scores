import json
import os
from io import BytesIO

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE_URL = "https://api.football-data.org/v4"
COMPETITION_CODE = "WC"
SEASON = 2026

STAGE_ORDER = [
    "Group Stage",
    "Round of 32",
    "Round of 16",
    "Quarter-finals",
    "Semi-finals",
    "Final"
]


DEFAULT_PARTICIPANTS = [
    {
        "name": "Ale",
        "categories": {
            "Category 1": [
                {
                    "id": 6,
                    "name": "Brasil"
                },
                {
                    "id": 30,
                    "name": "Portugal"
                }
            ],
            "Category 2": [
                {
                    "id": 34,
                    "name": "Colombia"
                },
                {
                    "id": 0,
                    "name": "Croacia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Noruega"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 0,
                    "name": "Costa de Marfil"
                },
                {
                    "id": 27,
                    "name": "México"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "República Democrática del Congo"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Nueva Zelanda"
                },
                {
                    "id": 0,
                    "name": "Haití"
                }
            ]
        }
    },
    {
        "name": "Isidro",
        "categories": {
            "Category 1": [
                {
                    "id": 29,
                    "name": "Francia"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 32,
                    "name": "Uruguay"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Senegal"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Chequia"
                },
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Irak"
                },
                {
                    "id": 0,
                    "name": "Sudáfrica"
                }
            ]
        }
    },
    {
        "name": "Fernando",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 30,
                    "name": "Portugal"
                }
            ],
            "Category 2": [
                {
                    "id": 33,
                    "name": "Países Bajos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Ecuador"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Túnez"
                },
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Nueva Zelanda"
                }
            ]
        }
    },
    {
        "name": "Claudia",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 26,
                    "name": "Argentina"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Ecuador"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "Catar"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Haití"
                }
            ]
        }
    },
    {
        "name": "Jorge Cañas [Coco]",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 29,
                    "name": "Francia"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Bélgica"
                },
                {
                    "id": 0,
                    "name": "Marruecos"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Noruega"
                },
                {
                    "id": 0,
                    "name": "Turquía"
                }
            ],
            "Category 4": [
                {
                    "id": 0,
                    "name": "Egipto"
                },
                {
                    "id": 0,
                    "name": "Costa de Marfil"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Ghana"
                },
                {
                    "id": 0,
                    "name": "Bosnia y Herzegovina"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Irak"
                },
                {
                    "id": 0,
                    "name": "Sudáfrica"
                }
            ]
        }
    },
    {
        "name": "Freddy",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 26,
                    "name": "Argentina"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Noruega"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Chequia"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Nueva Zelanda"
                }
            ]
        }
    },
    {
        "name": "Daniel",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 33,
                    "name": "Países Bajos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Ecuador"
                }
            ],
            "Category 4": [
                {
                    "id": 0,
                    "name": "Paraguay"
                },
                {
                    "id": 27,
                    "name": "México"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Bosnia y Herzegovina"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Nueva Zelanda"
                },
                {
                    "id": 0,
                    "name": "Haití"
                }
            ]
        }
    },
    {
        "name": "Gera F",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 26,
                    "name": "Argentina"
                }
            ],
            "Category 2": [
                {
                    "id": 33,
                    "name": "Países Bajos"
                },
                {
                    "id": 0,
                    "name": "Marruecos"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Noruega"
                }
            ],
            "Category 4": [
                {
                    "id": 0,
                    "name": "Paraguay"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Irán"
                },
                {
                    "id": 0,
                    "name": "Bosnia y Herzegovina"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Cabo Verde"
                }
            ]
        }
    },
    {
        "name": "Sebastian Flores",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 33,
                    "name": "Países Bajos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 0,
                    "name": "Ecuador"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Ghana"
                },
                {
                    "id": 0,
                    "name": "República Democrática del Congo"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Sudáfrica"
                },
                {
                    "id": 0,
                    "name": "Cabo Verde"
                }
            ]
        }
    },
    {
        "name": "Melani",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 29,
                    "name": "Francia"
                }
            ],
            "Category 2": [
                {
                    "id": 33,
                    "name": "Países Bajos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Noruega"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Australia"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Irak"
                },
                {
                    "id": 0,
                    "name": "Curazao"
                }
            ]
        }
    },
    {
        "name": "Jaz 💕",
        "categories": {
            "Category 1": [
                {
                    "id": 26,
                    "name": "Argentina"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Noruega"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Australia"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "República Democrática del Congo"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Uzbekistán"
                },
                {
                    "id": 0,
                    "name": "Panamá"
                }
            ]
        }
    },
    {
        "name": "Bere",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Curazao"
                }
            ]
        }
    },
    {
        "name": "Gaby",
        "categories": {
            "Category 1": [
                {
                    "id": 10,
                    "name": "Inglaterra"
                },
                {
                    "id": 6,
                    "name": "Brasil"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Corea del Sur"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Chequia"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Nueva Zelanda"
                },
                {
                    "id": 0,
                    "name": "Curazao"
                }
            ]
        }
    },
    {
        "name": "Bedy",
        "categories": {
            "Category 1": [
                {
                    "id": 31,
                    "name": "España"
                },
                {
                    "id": 26,
                    "name": "Argentina"
                }
            ],
            "Category 2": [
                {
                    "id": 0,
                    "name": "Marruecos"
                },
                {
                    "id": 34,
                    "name": "Colombia"
                }
            ],
            "Category 3": [
                {
                    "id": 0,
                    "name": "Japón"
                },
                {
                    "id": 28,
                    "name": "Estados Unidos"
                }
            ],
            "Category 4": [
                {
                    "id": 27,
                    "name": "México"
                },
                {
                    "id": 0,
                    "name": "Canadá"
                }
            ],
            "Category 5": [
                {
                    "id": 0,
                    "name": "Arabia Saudita"
                },
                {
                    "id": 0,
                    "name": "Escocia"
                }
            ],
            "Category 6": [
                {
                    "id": 0,
                    "name": "Panamá"
                },
                {
                    "id": 0,
                    "name": "Sudáfrica"
                }
            ]
        }
    }
]


def normalize_round(stage):
    if not stage:
        return ""

    stage_text = str(stage).upper().replace(" ", "_").replace("-", "_")

    if "GROUP" in stage_text:
        return "Group Stage"
    if "LAST_32" in stage_text or "ROUND_OF_32" in stage_text:
        return "Round of 32"
    if "LAST_16" in stage_text or "ROUND_OF_16" in stage_text:
        return "Round of 16"
    if "QUARTER" in stage_text:
        return "Quarter-finals"
    if "SEMI" in stage_text:
        return "Semi-finals"
    if "FINAL" in stage_text:
        return "Final"

    return str(stage)


TEAM_NAME_ALIASES = {
    "brasil": "brazil",
    "espana": "spain",
    "inglaterra": "england",
    "paises bajos": "netherlands",
    "alemania": "germany",
    "francia": "france",
    "portugal": "portugal",
    "argentina": "argentina",
    "colombia": "colombia",
    "uruguay": "uruguay",
    "croacia": "croatia",
    "marruecos": "morocco",
    "belgica": "belgium",
    "japon": "japan",
    "noruega": "norway",
    "turquia": "turkey",
    "senegal": "senegal",
    "ecuador": "ecuador",
    "estados unidos": "united states",
    "usa": "united states",
    "mexico": "mexico",
    "canada": "canada",
    "australia": "australia",
    "corea del sur": "south korea",
    "egipto": "egypt",
    "costa de marfil": "ivory coast",
    "arabia saudita": "saudi arabia",
    "republica democratica del congo": "dr congo",
    "rd congo": "dr congo",
    "chequia": "czech republic",
    "tunez": "tunisia",
    "catar": "qatar",
    "ghana": "ghana",
    "bosnia y herzegovina": "bosnia and herzegovina",
    "bosnia herzegovina": "bosnia and herzegovina",
    "escocia": "scotland",
    "iran": "iran",
    "nueva zelanda": "new zealand",
    "haiti": "haiti",
    "irak": "iraq",
    "sudafrica": "south africa",
    "panama": "panama",
    "cabo verde": "cape verde",
    "curazao": "curacao",
    "uzbekistan": "uzbekistan",
}


def normalize_team_name(name):
    text = str(name or "").strip().lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n",
        "Á": "a", "É": "e", "Í": "i", "Ó": "o", "Ú": "u", "Ü": "u", "Ñ": "n",
        "’": "'", ".": "", ",": ""
    }
    for old, new_value in replacements.items():
        text = text.replace(old, new_value)
    text = " ".join(text.split())
    return TEAM_NAME_ALIASES.get(text, text)


def teams_match(selected_team, api_team):
    selected_name = normalize_team_name(selected_team.get("name", ""))
    api_name = normalize_team_name(api_team.get("name", ""))
    short_name = normalize_team_name(api_team.get("shortName", ""))
    tla = normalize_team_name(api_team.get("tla", ""))

    return selected_name in {api_name, short_name, tla}


def is_finished(match):
    return match.get("status") == "FINISHED"


@st.cache_data(ttl=3600, show_spinner=False)
def get_worldcup_fixtures(api_key):
    url = f"{BASE_URL}/competitions/{COMPETITION_CODE}/matches"

    headers = {
        "X-Auth-Token": api_key
    }

    params = {
        "season": SEASON
    }

    response = requests.get(url, headers=headers, params=params, timeout=30)
    response.raise_for_status()

    data = response.json()
    return data.get("matches", [])


def match_winner_side(match):
    score = match.get("score", {}) or {}
    winner = score.get("winner")

    if winner in {"HOME_TEAM", "AWAY_TEAM", "DRAW"}:
        return winner

    # Fallback for APIs/data snapshots where winner is missing but final score exists.
    full_time = score.get("fullTime", {}) or {}
    home_score = full_time.get("home")
    away_score = full_time.get("away")

    if home_score is None or away_score is None:
        return None
    if home_score > away_score:
        return "HOME_TEAM"
    if away_score > home_score:
        return "AWAY_TEAM"
    return "DRAW"


def calculate_team_points(team, fixtures):
    points = 0
    wins = 0
    draws = 0
    advanced_stages = 0
    champion_bonus = 0

    reached_stages = set()

    for match in fixtures:
        if not is_finished(match):
            continue

        home = match.get("homeTeam", {}) or {}
        away = match.get("awayTeam", {}) or {}
        round_name = normalize_round(match.get("stage", ""))

        is_home_team = teams_match(team, home)
        is_away_team = teams_match(team, away)

        if not is_home_team and not is_away_team:
            continue

        reached_stages.add(round_name)
        winner = match_winner_side(match)

        if (is_home_team and winner == "HOME_TEAM") or (is_away_team and winner == "AWAY_TEAM"):
            points += 3
            wins += 1
        elif winner == "DRAW":
            points += 1
            draws += 1

        if round_name == "Final" and (
            (is_home_team and winner == "HOME_TEAM") or
            (is_away_team and winner == "AWAY_TEAM")
        ):
            points += 10
            champion_bonus = 10

    for stage in STAGE_ORDER[1:]:
        if stage in reached_stages:
            points += 2
            advanced_stages += 1

    reached_stages_sorted = sorted(
        reached_stages,
        key=lambda stage: STAGE_ORDER.index(stage) if stage in STAGE_ORDER else 99
    )

    return {
        "points": points,
        "wins": wins,
        "draws": draws,
        "advanced_stages": advanced_stages,
        "champion_bonus": champion_bonus,
        "reached_stages": ", ".join(reached_stages_sorted)
    }

def build_leaderboard(participants, fixtures):
    leaderboard = []

    for participant in participants:
        total_points = 0
        team_breakdown = []

        for category_name, teams in participant["categories"].items():
            for team in teams:
                result = calculate_team_points(team, fixtures)
                total_points += result["points"]

                team_breakdown.append({
                    "Participant": participant["name"],
                    "Category": category_name,
                    "Team": team["name"],
                    "Team ID": team["id"],
                    "Team Points": result["points"],
                    "Wins": result["wins"],
                    "Draws": result["draws"],
                    "Advanced Stages": result["advanced_stages"],
                    "Champion Bonus": result["champion_bonus"],
                    "Reached Stages": result["reached_stages"]
                })

        leaderboard.append({
            "Participant": participant["name"],
            "Total Points": total_points,
            "Teams": team_breakdown
        })

    leaderboard.sort(key=lambda x: x["Total Points"], reverse=True)

    return leaderboard


def create_dataframes(leaderboard):
    leaderboard_rows = []

    for index, entry in enumerate(leaderboard, start=1):
        leaderboard_rows.append({
            "Rank": index,
            "Participant": entry["Participant"],
            "Total Points": entry["Total Points"]
        })

    details_rows = []

    for entry in leaderboard:
        for team in entry["Teams"]:
            details_rows.append(team)

    category_rows = []

    for entry in leaderboard:
        category_totals = {}

        for team in entry["Teams"]:
            category = team["Category"]

            if category not in category_totals:
                category_totals[category] = 0

            category_totals[category] += team["Team Points"]

        for category, points in category_totals.items():
            category_rows.append({
                "Participant": entry["Participant"],
                "Category": category,
                "Category Points": points
            })

    leaderboard_df = pd.DataFrame(leaderboard_rows)
    details_df = pd.DataFrame(details_rows)
    category_df = pd.DataFrame(category_rows)

    return leaderboard_df, details_df, category_df


def style_header(row):
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def add_borders(sheet):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def create_excel_file(leaderboard_df, details_df, category_df):
    workbook = Workbook()

    leaderboard_sheet = workbook.active
    leaderboard_sheet.title = "Leaderboard"

    leaderboard_sheet.append(list(leaderboard_df.columns))

    for row in leaderboard_df.itertuples(index=False):
        leaderboard_sheet.append(list(row))

    style_header(leaderboard_sheet[1])
    add_borders(leaderboard_sheet)

    leaderboard_sheet.column_dimensions["A"].width = 10
    leaderboard_sheet.column_dimensions["B"].width = 25
    leaderboard_sheet.column_dimensions["C"].width = 15

    details_sheet = workbook.create_sheet("Team Details")
    details_sheet.append(list(details_df.columns))

    for row in details_df.itertuples(index=False):
        details_sheet.append(list(row))

    style_header(details_sheet[1])
    add_borders(details_sheet)

    details_widths = {
        "A": 25,
        "B": 18,
        "C": 25,
        "D": 12,
        "E": 15,
        "F": 10,
        "G": 10,
        "H": 18,
        "I": 18,
        "J": 60
    }

    for column, width in details_widths.items():
        details_sheet.column_dimensions[column].width = width

    category_sheet = workbook.create_sheet("Category Summary")
    category_sheet.append(list(category_df.columns))

    for row in category_df.itertuples(index=False):
        category_sheet.append(list(row))

    style_header(category_sheet[1])
    add_borders(category_sheet)

    category_sheet.column_dimensions["A"].width = 25
    category_sheet.column_dimensions["B"].width = 18
    category_sheet.column_dimensions["C"].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def validate_participants(participants):
    for participant in participants:
        if "name" not in participant:
            raise ValueError("Each participant needs a name.")

        if "categories" not in participant:
            raise ValueError(f"{participant['name']} needs categories.")

        if len(participant["categories"]) != 6:
            raise ValueError(f"{participant['name']} must have exactly 6 categories.")

        for category_name, teams in participant["categories"].items():
            if len(teams) != 2:
                raise ValueError(
                    f"{participant['name']} - {category_name} must have exactly 2 teams."
                )

            for team in teams:
                if "id" not in team or "name" not in team:
                    raise ValueError(
                        f"Each team in {participant['name']} - {category_name} needs an id and name."
                    )


def get_configured_api_key():
    """Return an API token from Streamlit secrets, environment variables, or user input."""
    token = ""

    try:
        token = st.secrets.get("FOOTBALL_DATA_API_TOKEN", "")
    except Exception:
        token = ""

    if not token:
        token = os.getenv("FOOTBALL_DATA_API_TOKEN", "")

    return token


def render_results(leaderboard_df, category_df, details_df, excel_file):
    st.success("Points calculated successfully.")

    metric_cols = st.columns(3)
    metric_cols[0].metric("Participants", len(leaderboard_df))
    metric_cols[1].metric("Teams Tracked", len(details_df))
    metric_cols[2].metric("Top Score", int(leaderboard_df["Total Points"].max()) if not leaderboard_df.empty else 0)

    st.subheader("Leaderboard")
    st.dataframe(leaderboard_df, use_container_width=True, hide_index=True)

    st.subheader("Category Summary")
    st.dataframe(category_df, use_container_width=True, hide_index=True)

    st.subheader("Team Details")
    st.dataframe(details_df, use_container_width=True, hide_index=True)

    st.download_button(
        label="Download Excel File",
        data=excel_file,
        file_name="worldcup_points.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def main():
    st.set_page_config(
        page_title="World Cup Points App",
        page_icon="🏆",
        layout="wide"
    )

    st.title("🏆 FIFA World Cup Points App")
    st.write(
        "Calculate participant points using football-data.org data and export the results to Excel."
    )

    configured_api_key = get_configured_api_key()

    with st.sidebar:
        st.header("Settings")

        manual_api_key = st.text_input(
            "football-data.org API Token",
            value="" if configured_api_key else "",
            type="password",
            placeholder="Paste your API key here"
        )

        api_key = manual_api_key.strip() or configured_api_key

        if configured_api_key and not manual_api_key:
            st.success("Using API token from app secrets/environment.")

        st.write("Competition:")
        st.write(f"Competition Code: `{COMPETITION_CODE}`")
        st.write(f"Season: `{SEASON}`")

        st.info(
            "Scoring: win = 3 points, draw = 1 point, each advanced phase = 2 points, World Cup champion = 10 points."
        )

        refresh_cache = st.button("Refresh API Cache")
        if refresh_cache:
            get_worldcup_fixtures.clear()
            st.success("API cache cleared.")

    st.subheader("Participants")
    st.write(
        "Edit the JSON below. Each participant should have 6 categories, and each category should have 2 teams."
    )

    default_json = json.dumps(DEFAULT_PARTICIPANTS, indent=4, ensure_ascii=False)
    participants_json = st.text_area(
        "Participants JSON",
        value=st.session_state.get("participants_json", default_json),
        height=500
    )
    st.session_state["participants_json"] = participants_json

    calculate_button = st.button("Calculate Points", type="primary")

    if not calculate_button:
        st.caption("Enter an API token and click Calculate Points to load the latest fixtures.")
        return

    if not api_key:
        st.error(
            "Please enter a football-data.org API token, or set FOOTBALL_DATA_API_TOKEN in Streamlit secrets/environment variables."
        )
        return

    try:
        participants = json.loads(participants_json)
        validate_participants(participants)

        with st.spinner("Getting World Cup fixtures and calculating points..."):
            fixtures = get_worldcup_fixtures(api_key)
            leaderboard = build_leaderboard(participants, fixtures)
            leaderboard_df, details_df, category_df = create_dataframes(leaderboard)
            excel_file = create_excel_file(leaderboard_df, details_df, category_df)

        render_results(leaderboard_df, category_df, details_df, excel_file)

    except json.JSONDecodeError as error:
        st.error(f"The participants section is not valid JSON: {error}")
    except requests.exceptions.Timeout:
        st.error("The football-data.org request timed out. Try again in a moment.")
    except requests.exceptions.HTTPError as error:
        response = getattr(error, "response", None)
        status_code = response.status_code if response is not None else "unknown"
        message = response.text if response is not None else str(error)
        st.error(f"API request failed with status {status_code}: {message}")
    except requests.exceptions.RequestException as error:
        st.error(f"Could not connect to football-data.org: {error}")
    except ValueError as error:
        st.error(str(error))
    except Exception as error:
        st.error(f"Unexpected error: {error}")


if __name__ == "__main__":
    main()
